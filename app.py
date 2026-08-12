
from flask import Flask, request, redirect, url_for, session, render_template, flash, jsonify, send_file
import sqlite3, os, uuid, datetime, re, io, mimetypes
from urllib.parse import quote
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename
from openpyxl import load_workbook
import xlrd

APP_DIR = os.path.dirname(os.path.abspath(__file__))
DATABASE_URL = os.environ.get("DATABASE_URL", "").strip()
IS_POSTGRES = DATABASE_URL.startswith("postgres://") or DATABASE_URL.startswith("postgresql://")
DB_PATH = os.path.join(APP_DIR, "mini_crm.db")
UPLOAD_DIR = os.path.join(APP_DIR, "uploads")
EMPLOYEE_UPLOAD_DIR = os.path.join(UPLOAD_DIR, "employees")
os.makedirs(UPLOAD_DIR, exist_ok=True)
os.makedirs(EMPLOYEE_UPLOAD_DIR, exist_ok=True)

ALLOWED_DOC_EXTENSIONS = {"pdf","jpg","jpeg","png","webp","doc","docx"}
ALLOWED_PHOTO_EXTENSIONS = {"jpg","jpeg","png","webp"}

def allowed_file(filename, allowed):
    return "." in filename and filename.rsplit(".",1)[1].lower() in allowed

app = Flask(__name__)
app.secret_key = os.environ.get("GAUR_CRM_SECRET", "GAUR-CRM-LOCAL-2026-CHANGE-ME")
app.config["MAX_CONTENT_LENGTH"] = int(os.environ.get("MAX_UPLOAD_MB", "20")) * 1024 * 1024
if os.environ.get("CLOUD_MODE", "").lower() in ("1","true","yes") or os.environ.get("RAILWAY_ENVIRONMENT"):
    app.config.update(
        SESSION_COOKIE_HTTPONLY=True,
        SESSION_COOKIE_SAMESITE="Lax",
        SESSION_COOKIE_SECURE=True,
    )

class CursorProxy:
    def __init__(self, cur, lastrowid=None, buffered_row=None):
        self.cur=cur
        self.lastrowid=lastrowid
        self._buffered_row=buffered_row
    def fetchone(self):
        if self._buffered_row is not None:
            r=self._buffered_row; self._buffered_row=None; return r
        return self.cur.fetchone()
    def fetchall(self):
        rows=[]
        if self._buffered_row is not None:
            rows.append(self._buffered_row); self._buffered_row=None
        rows.extend(self.cur.fetchall())
        return rows

class PostgresConnection:
    def __init__(self, url):
        import psycopg
        from psycopg.rows import dict_row
        self.con=psycopg.connect(url, row_factory=dict_row, connect_timeout=15)
    def _sql(self, sql):
        # App code uses SQLite-style ? placeholders; psycopg uses %s.
        sql=sql.replace("?", "%s")
        sql=re.sub(r"INSERT\s+OR\s+IGNORE\s+INTO\s+companies\(code,name\)\s+VALUES\(%s,%s\)",
                   "INSERT INTO companies(code,name) VALUES(%s,%s) ON CONFLICT (code) DO NOTHING", sql, flags=re.I)
        return sql
    def execute(self, sql, params=()):
        sql2=self._sql(sql)
        cur=self.con.cursor()
        # Preserve SQLite's cursor.lastrowid behavior used by Employee creation.
        need_id=bool(re.match(r"\s*INSERT\s+INTO\s+employee_master\b", sql2, re.I)) and "RETURNING" not in sql2.upper()
        if need_id:
            sql2=sql2.rstrip().rstrip(';')+" RETURNING id"
        cur.execute(sql2, params or ())
        if need_id:
            row=cur.fetchone()
            rid=row["id"] if row else None
            return CursorProxy(cur,lastrowid=rid)
        return CursorProxy(cur)
    def cursor(self):
        return self
    def executescript(self, script):
        cur=self.con.cursor()
        for stmt in script.split(';'):
            if stmt.strip(): cur.execute(stmt)
        return CursorProxy(cur)
    def commit(self): self.con.commit()
    def rollback(self): self.con.rollback()
    def close(self): self.con.close()

def db():
    if IS_POSTGRES:
        return PostgresConnection(DATABASE_URL)
    con = sqlite3.connect(DB_PATH, timeout=30)
    con.row_factory = sqlite3.Row
    try:
        con.execute("PRAGMA journal_mode=WAL")
        con.execute("PRAGMA synchronous=NORMAL")
        con.execute("PRAGMA busy_timeout=30000")
    except Exception:
        pass
    return con

def init_db():
    con=db(); cur=con.cursor()
    if IS_POSTGRES:
        cur.executescript("""
        CREATE TABLE IF NOT EXISTS companies(
          id BIGSERIAL PRIMARY KEY, code TEXT UNIQUE NOT NULL, name TEXT NOT NULL
        );
        CREATE TABLE IF NOT EXISTS users(
          id BIGSERIAL PRIMARY KEY, login_id TEXT UNIQUE NOT NULL, password_hash TEXT NOT NULL,
          full_name TEXT NOT NULL, role TEXT NOT NULL, company_code TEXT, active INTEGER DEFAULT 1
        );
        CREATE TABLE IF NOT EXISTS visitors(
          id BIGSERIAL PRIMARY KEY, visitor_id TEXT UNIQUE, company_code TEXT NOT NULL, visit_date TEXT,
          client_name TEXT, mobile TEXT, email TEXT, city TEXT, country TEXT, visa_type TEXT,
          coordinating_with TEXT, purpose TEXT, status TEXT, remarks TEXT, created_by TEXT
        );
        CREATE TABLE IF NOT EXISTS leads(
          id BIGSERIAL PRIMARY KEY, lead_id TEXT UNIQUE, company_code TEXT NOT NULL, client_name TEXT,
          mobile TEXT, email TEXT, city TEXT, country TEXT, visa_type TEXT, source TEXT,
          duplicate_type TEXT DEFAULT '', assigned_am BIGINT, status TEXT DEFAULT 'New',
          followup_date TEXT DEFAULT '', remarks TEXT DEFAULT '', imported_by TEXT, imported_at TEXT,
          assigned_at TEXT DEFAULT '', assigned_by TEXT DEFAULT ''
        );
        CREATE TABLE IF NOT EXISTS imports(
          id BIGSERIAL PRIMARY KEY, company_code TEXT, filename TEXT, total_rows INTEGER, new_rows INTEGER,
          same_company_duplicates INTEGER, cross_company_duplicates INTEGER, invalid_mobiles INTEGER,
          imported_by TEXT, imported_at TEXT
        );
        CREATE TABLE IF NOT EXISTS employee_master(
          id BIGSERIAL PRIMARY KEY, employee_code TEXT UNIQUE NOT NULL, company_code TEXT NOT NULL,
          full_name TEXT NOT NULL, father_spouse_name TEXT DEFAULT '', designation TEXT DEFAULT '',
          department TEXT DEFAULT '', mobile TEXT DEFAULT '', alternate_mobile TEXT DEFAULT '', email TEXT DEFAULT '',
          date_of_birth TEXT DEFAULT '', joining_date TEXT DEFAULT '', address TEXT DEFAULT '', city TEXT DEFAULT '',
          state TEXT DEFAULT '', pin_code TEXT DEFAULT '', emergency_contact_name TEXT DEFAULT '',
          emergency_contact_number TEXT DEFAULT '', aadhaar_last4 TEXT DEFAULT '', pan_number TEXT DEFAULT '',
          bank_name TEXT DEFAULT '', account_last4 TEXT DEFAULT '', ifsc_code TEXT DEFAULT '',
          salary_reference TEXT DEFAULT '', employment_status TEXT DEFAULT 'Active', remarks TEXT DEFAULT '',
          photo_path TEXT DEFAULT '', photo_data BYTEA, photo_mime TEXT DEFAULT '', created_by TEXT DEFAULT '',
          created_at TEXT DEFAULT '', updated_at TEXT DEFAULT ''
        );
        CREATE TABLE IF NOT EXISTS employee_documents(
          id BIGSERIAL PRIMARY KEY, employee_id BIGINT NOT NULL, document_type TEXT NOT NULL,
          document_name TEXT DEFAULT '', file_name TEXT NOT NULL, stored_path TEXT NOT NULL,
          file_data BYTEA, mime_type TEXT DEFAULT '', uploaded_by TEXT DEFAULT '', uploaded_at TEXT DEFAULT '',
          FOREIGN KEY(employee_id) REFERENCES employee_master(id)
        );
        CREATE TABLE IF NOT EXISTS employee_performance(
          id BIGSERIAL PRIMARY KEY, employee_id BIGINT NOT NULL, performance_date TEXT NOT NULL,
          leads_assigned INTEGER DEFAULT 0, calls_done INTEGER DEFAULT 0, followups_done INTEGER DEFAULT 0,
          office_visits INTEGER DEFAULT 0, interested_clients INTEGER DEFAULT 0, enrollments INTEGER DEFAULT 0,
          collection_amount REAL DEFAULT 0, target_amount REAL DEFAULT 0, attendance_status TEXT DEFAULT 'Present',
          manager_rating REAL DEFAULT 0, remarks TEXT DEFAULT '', created_by TEXT DEFAULT '', created_at TEXT DEFAULT '',
          FOREIGN KEY(employee_id) REFERENCES employee_master(id)
        );
        CREATE TABLE IF NOT EXISTS allocation_history(
          id BIGSERIAL PRIMARY KEY, company_code TEXT NOT NULL, am_user_id BIGINT NOT NULL,
          quantity INTEGER DEFAULT 0, allocated_by TEXT DEFAULT '', allocated_at TEXT DEFAULT ''
        );
        CREATE TABLE IF NOT EXISTS client_cases(
          id BIGSERIAL PRIMARY KEY, case_id TEXT UNIQUE NOT NULL, company_code TEXT NOT NULL,
          client_name TEXT NOT NULL, mobile TEXT DEFAULT '', country TEXT DEFAULT '', visa_type TEXT DEFAULT '',
          enrollment_date TEXT DEFAULT '', booking_amount REAL DEFAULT 0, second_payment REAL DEFAULT 0,
          total_received REAL DEFAULT 0, payment_status TEXT DEFAULT 'Pending',
          filing_status TEXT DEFAULT 'Documents Pending', assigned_employee_id BIGINT, remarks TEXT DEFAULT '',
          created_by TEXT DEFAULT '', created_at TEXT DEFAULT '', updated_at TEXT DEFAULT ''
        );
        CREATE INDEX IF NOT EXISTS idx_users_role_company ON users(role,company_code);
        CREATE INDEX IF NOT EXISTS idx_visitors_company ON visitors(company_code,id);
        CREATE INDEX IF NOT EXISTS idx_leads_company_am ON leads(company_code,assigned_am,id);
        CREATE INDEX IF NOT EXISTS idx_cases_company_employee ON client_cases(company_code,assigned_employee_id,id);
        """)
        # Cloud-safe migrations for deployments upgraded in place.
        cur.execute("ALTER TABLE employee_master ADD COLUMN IF NOT EXISTS photo_data BYTEA")
        cur.execute("ALTER TABLE employee_master ADD COLUMN IF NOT EXISTS photo_mime TEXT DEFAULT ''")
        cur.execute("ALTER TABLE employee_documents ADD COLUMN IF NOT EXISTS file_data BYTEA")
        cur.execute("ALTER TABLE employee_documents ADD COLUMN IF NOT EXISTS mime_type TEXT DEFAULT ''")
        cur.execute("ALTER TABLE leads ADD COLUMN IF NOT EXISTS assigned_at TEXT DEFAULT ''")
        cur.execute("ALTER TABLE leads ADD COLUMN IF NOT EXISTS assigned_by TEXT DEFAULT ''")
    else:
        cur.executescript("""
        CREATE TABLE IF NOT EXISTS companies(id INTEGER PRIMARY KEY AUTOINCREMENT,code TEXT UNIQUE NOT NULL,name TEXT NOT NULL);
        CREATE TABLE IF NOT EXISTS users(id INTEGER PRIMARY KEY AUTOINCREMENT,login_id TEXT UNIQUE NOT NULL,password_hash TEXT NOT NULL,full_name TEXT NOT NULL,role TEXT NOT NULL,company_code TEXT,active INTEGER DEFAULT 1);
        CREATE TABLE IF NOT EXISTS visitors(id INTEGER PRIMARY KEY AUTOINCREMENT,visitor_id TEXT UNIQUE,company_code TEXT NOT NULL,visit_date TEXT,client_name TEXT,mobile TEXT,email TEXT,city TEXT,country TEXT,visa_type TEXT,coordinating_with TEXT,purpose TEXT,status TEXT,remarks TEXT,created_by TEXT);
        CREATE TABLE IF NOT EXISTS leads(id INTEGER PRIMARY KEY AUTOINCREMENT,lead_id TEXT UNIQUE,company_code TEXT NOT NULL,client_name TEXT,mobile TEXT,email TEXT,city TEXT,country TEXT,visa_type TEXT,source TEXT,duplicate_type TEXT DEFAULT '',assigned_am INTEGER,status TEXT DEFAULT 'New',followup_date TEXT DEFAULT '',remarks TEXT DEFAULT '',imported_by TEXT,imported_at TEXT,assigned_at TEXT DEFAULT '',assigned_by TEXT DEFAULT '');
        CREATE TABLE IF NOT EXISTS imports(id INTEGER PRIMARY KEY AUTOINCREMENT,company_code TEXT,filename TEXT,total_rows INTEGER,new_rows INTEGER,same_company_duplicates INTEGER,cross_company_duplicates INTEGER,invalid_mobiles INTEGER,imported_by TEXT,imported_at TEXT);
        CREATE TABLE IF NOT EXISTS employee_master(id INTEGER PRIMARY KEY AUTOINCREMENT,employee_code TEXT UNIQUE NOT NULL,company_code TEXT NOT NULL,full_name TEXT NOT NULL,father_spouse_name TEXT DEFAULT '',designation TEXT DEFAULT '',department TEXT DEFAULT '',mobile TEXT DEFAULT '',alternate_mobile TEXT DEFAULT '',email TEXT DEFAULT '',date_of_birth TEXT DEFAULT '',joining_date TEXT DEFAULT '',address TEXT DEFAULT '',city TEXT DEFAULT '',state TEXT DEFAULT '',pin_code TEXT DEFAULT '',emergency_contact_name TEXT DEFAULT '',emergency_contact_number TEXT DEFAULT '',aadhaar_last4 TEXT DEFAULT '',pan_number TEXT DEFAULT '',bank_name TEXT DEFAULT '',account_last4 TEXT DEFAULT '',ifsc_code TEXT DEFAULT '',salary_reference TEXT DEFAULT '',employment_status TEXT DEFAULT 'Active',remarks TEXT DEFAULT '',photo_path TEXT DEFAULT '',photo_data BLOB,photo_mime TEXT DEFAULT '',created_by TEXT DEFAULT '',created_at TEXT DEFAULT '',updated_at TEXT DEFAULT '');
        CREATE TABLE IF NOT EXISTS employee_documents(id INTEGER PRIMARY KEY AUTOINCREMENT,employee_id INTEGER NOT NULL,document_type TEXT NOT NULL,document_name TEXT DEFAULT '',file_name TEXT NOT NULL,stored_path TEXT NOT NULL,file_data BLOB,mime_type TEXT DEFAULT '',uploaded_by TEXT DEFAULT '',uploaded_at TEXT DEFAULT '',FOREIGN KEY(employee_id) REFERENCES employee_master(id));
        CREATE TABLE IF NOT EXISTS employee_performance(id INTEGER PRIMARY KEY AUTOINCREMENT,employee_id INTEGER NOT NULL,performance_date TEXT NOT NULL,leads_assigned INTEGER DEFAULT 0,calls_done INTEGER DEFAULT 0,followups_done INTEGER DEFAULT 0,office_visits INTEGER DEFAULT 0,interested_clients INTEGER DEFAULT 0,enrollments INTEGER DEFAULT 0,collection_amount REAL DEFAULT 0,target_amount REAL DEFAULT 0,attendance_status TEXT DEFAULT 'Present',manager_rating REAL DEFAULT 0,remarks TEXT DEFAULT '',created_by TEXT DEFAULT '',created_at TEXT DEFAULT '',FOREIGN KEY(employee_id) REFERENCES employee_master(id));
        CREATE TABLE IF NOT EXISTS allocation_history(id INTEGER PRIMARY KEY AUTOINCREMENT,company_code TEXT NOT NULL,am_user_id INTEGER NOT NULL,quantity INTEGER DEFAULT 0,allocated_by TEXT DEFAULT '',allocated_at TEXT DEFAULT '');
        CREATE TABLE IF NOT EXISTS client_cases(id INTEGER PRIMARY KEY AUTOINCREMENT,case_id TEXT UNIQUE NOT NULL,company_code TEXT NOT NULL,client_name TEXT NOT NULL,mobile TEXT DEFAULT '',country TEXT DEFAULT '',visa_type TEXT DEFAULT '',enrollment_date TEXT DEFAULT '',booking_amount REAL DEFAULT 0,second_payment REAL DEFAULT 0,total_received REAL DEFAULT 0,payment_status TEXT DEFAULT 'Pending',filing_status TEXT DEFAULT 'Documents Pending',assigned_employee_id INTEGER,remarks TEXT DEFAULT '',created_by TEXT DEFAULT '',created_at TEXT DEFAULT '',updated_at TEXT DEFAULT '');
        """)
        cols=[r[1] for r in cur.execute("PRAGMA table_info(employee_master)").fetchall()]
        for col,ddl in [("photo_path","TEXT DEFAULT ''"),("photo_data","BLOB"),("photo_mime","TEXT DEFAULT ''")]:
            if col not in cols: cur.execute(f"ALTER TABLE employee_master ADD COLUMN {col} {ddl}")
        dcols=[r[1] for r in cur.execute("PRAGMA table_info(employee_documents)").fetchall()]
        for col,ddl in [("file_data","BLOB"),("mime_type","TEXT DEFAULT ''")]:
            if col not in dcols: cur.execute(f"ALTER TABLE employee_documents ADD COLUMN {col} {ddl}")
        lead_cols=[r[1] for r in cur.execute("PRAGMA table_info(leads)").fetchall()]
        if "assigned_at" not in lead_cols: cur.execute("ALTER TABLE leads ADD COLUMN assigned_at TEXT DEFAULT ''")
        if "assigned_by" not in lead_cols: cur.execute("ALTER TABLE leads ADD COLUMN assigned_by TEXT DEFAULT ''")

    for code,name in [("SCIC","Smart Choice Immigration Consultants"),("WWIC","White Wave Immigration Consultants")]:
        if IS_POSTGRES:
            cur.execute("INSERT INTO companies(code,name) VALUES(?,?) ON CONFLICT (code) DO NOTHING",(code,name))
        else:
            cur.execute("INSERT OR IGNORE INTO companies(code,name) VALUES(?,?)",(code,name))

    defaults=[
      ("md.sandeep","Sandeep Gaur","MD",None,"Sandeep@2026"),
      ("gm.smartchoice","Smart Choice General Manager","GM","SCIC","SCIC@GM2026#"),
      ("gm.whitewave","White Wave General Manager","GM","WWIC","WWIC@GM2026#"),
      ("reception.scic","Smart Choice Receptionist","RECEPTION","SCIC","Welcome@123"),
      ("reception.wwic","White Wave Receptionist","RECEPTION","WWIC","Welcome@123"),
    ]
    legacy_map={"gm.scic":"gm.smartchoice","gm.wwic":"gm.whitewave"}
    for old_id,new_id in legacy_map.items():
        oldrow=cur.execute("SELECT id FROM users WHERE lower(login_id)=?",(old_id,)).fetchone()
        newrow=cur.execute("SELECT id FROM users WHERE lower(login_id)=?",(new_id,)).fetchone()
        if oldrow and not newrow: cur.execute("UPDATE users SET login_id=? WHERE id=?",(new_id,oldrow["id"]))

    for login_id,name,role,company,pw in defaults:
        existing=cur.execute("SELECT id FROM users WHERE lower(login_id)=lower(?)",(login_id,)).fetchone()
        if existing:
            # Preserve the user's current cloud password across redeploys.
            cur.execute("UPDATE users SET full_name=?,role=?,company_code=?,active=1 WHERE id=?",
                        (name,role,company,existing["id"]))
        else:
            cur.execute("INSERT INTO users(login_id,password_hash,full_name,role,company_code,active) VALUES(?,?,?,?,?,1)",
                        (login_id,generate_password_hash(pw),name,role,company))
    con.commit(); con.close()

def current_user():
    uid = session.get("uid")
    if not uid: return None
    con=db()
    u=con.execute("SELECT * FROM users WHERE id=? AND active=1",(uid,)).fetchone()
    con.close()
    return u

def require_roles(*roles):
    def deco(fn):
        from functools import wraps
        @wraps(fn)
        def wrap(*a, **kw):
            u=current_user()
            if not u: return redirect(url_for("login"))
            if roles and u["role"] not in roles:
                flash("Access denied","error")
                return redirect(url_for("dashboard"))
            return fn(*a,**kw)
        return wrap
    return deco

@app.route("/healthz")
def healthz():
    try:
        con=db(); con.execute("SELECT 1").fetchone(); con.close()
        return jsonify({"status":"ok","database":"postgresql" if IS_POSTGRES else "sqlite"}),200
    except Exception as exc:
        return jsonify({"status":"error","detail":str(exc)[:200]}),503

@app.route("/", methods=["GET","POST"])
def login():
    if current_user():
        return redirect(url_for("dashboard"))

    if request.method=="POST":
        lid=request.form.get("login_id","").strip().lower()
        pw=request.form.get("password","")
        con=db()
        u=con.execute("SELECT * FROM users WHERE lower(login_id)=? AND active=1",(lid,)).fetchone()
        con.close()

        if u and check_password_hash(u["password_hash"],pw):
            # No company or role is revealed before authentication.
            # Company/role access is resolved only AFTER valid credentials.
            session["uid"]=u["id"]
            return redirect(url_for("dashboard"))

        flash("Access denied. Invalid Login ID or password.","error")

    return render_template("login_gaur.html")

@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))


@app.route("/central-status")
def central_status():
    u=current_user()
    if not u:
        return redirect(url_for("login"))
    con=db()
    visitors=con.execute("SELECT COUNT(*) c FROM visitors").fetchone()["c"]
    ams=con.execute("SELECT COUNT(*) c FROM users WHERE role='AM'").fetchone()["c"]
    leads=con.execute("SELECT COUNT(*) c FROM leads").fetchone()["c"]
    con.close()
    return jsonify({
        "status":"online",
        "mode":"CENTRAL_SHARED_SERVER",
        "database":os.path.basename(DB_PATH),
        "visitors":visitors,
        "assistant_managers":ams,
        "leads":leads,
        "logged_in_as":u["login_id"],
        "company":u["company_code"] or "BOTH"
    })

@app.route("/dashboard")
@require_roles("MD","GM","RECEPTION","AM")
def dashboard():
    u=current_user(); con=db()
    where=[]; params=[]
    if u["role"]!="MD":
        where.append("company_code=?"); params.append(u["company_code"])
    if u["role"]=="AM":
        where.append("assigned_am=?"); params.append(u["id"])
    lead_where=(" WHERE "+" AND ".join(where)) if where else ""
    total_leads=con.execute("SELECT COUNT(*) c FROM leads"+lead_where,params).fetchone()["c"]
    allocated=con.execute("SELECT COUNT(*) c FROM leads"+lead_where+(" AND " if lead_where else " WHERE ")+"assigned_am IS NOT NULL",params).fetchone()["c"]
    interested=con.execute("SELECT COUNT(*) c FROM leads"+lead_where+(" AND " if lead_where else " WHERE ")+"status='Interested'",params).fetchone()["c"]
    unallocated=con.execute("SELECT COUNT(*) c FROM leads"+lead_where+(" AND " if lead_where else " WHERE ")+"assigned_am IS NULL",params).fetchone()["c"]
    cross_dups=con.execute("SELECT COUNT(*) c FROM leads"+lead_where+(" AND " if lead_where else " WHERE ")+"duplicate_type='cross'",params).fetchone()["c"]
    visitors_where="" if u["role"]=="MD" else " WHERE company_code=?"
    visitors_params=[] if u["role"]=="MD" else [u["company_code"]]
    visitors=con.execute("SELECT COUNT(*) c FROM visitors"+visitors_where,visitors_params).fetchone()["c"]
    recent=con.execute("SELECT * FROM leads"+lead_where+" ORDER BY id DESC LIMIT 8",params).fetchall()
    case_where="" if u["role"]=="MD" else " WHERE company_code=?"
    case_params=[] if u["role"]=="MD" else [u["company_code"]]
    cases=con.execute("SELECT COUNT(*) c FROM client_cases"+case_where,case_params).fetchone()["c"]
    revenue=con.execute("SELECT COALESCE(SUM(total_received),0) s FROM client_cases"+case_where,case_params).fetchone()["s"]
    con.close()
    return render_template("dashboard.html",u=u,total_leads=total_leads,allocated=allocated,interested=interested,
                           unallocated=unallocated,cross_dups=cross_dups,visitors=visitors,recent=recent,cases=cases,revenue=revenue)

@app.route("/reception", methods=["GET","POST"])
@require_roles("MD","GM","RECEPTION")
def reception():
    u=current_user()
    con=db()
    if request.method=="POST":
        company=request.form.get("company_code") if u["role"]=="MD" else u["company_code"]
        vid=f"{company}-VIS-{datetime.datetime.now().strftime('%Y%m%d%H%M%S')}"
        con.execute("""INSERT INTO visitors(visitor_id,company_code,visit_date,client_name,mobile,email,city,country,visa_type,coordinating_with,purpose,status,remarks,created_by)
                     VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                    (vid,company,request.form.get("visit_date"),request.form.get("client_name"),request.form.get("mobile"),
                     request.form.get("email"),request.form.get("city"),request.form.get("country"),request.form.get("visa_type"),
                     request.form.get("coordinating_with"),request.form.get("purpose"),request.form.get("status"),request.form.get("remarks"),u["login_id"]))
        con.commit()
        flash("Visitor saved","success")
    if u["role"]=="MD":
        rows=con.execute("SELECT * FROM visitors ORDER BY id DESC LIMIT 100").fetchall()
    else:
        rows=con.execute("SELECT * FROM visitors WHERE company_code=? ORDER BY id DESC LIMIT 100",(u["company_code"],)).fetchall()
    con.close()
    return render_template("reception.html",u=u,rows=rows)

def normalize_header(x):
    return re.sub(r'[^a-z0-9]+','',str(x or '').lower())

FIELD_ALIASES = {
    "client_name":{"name","clientname","fullname","leadname","customername"},
    "mobile":{"mobile","mobileno","mobilephone","phone","phoneno","contact","contactno","contactnumber"},
    "email":{"email","emailid","emailaddress"},
    "city":{"city","place","location","district"},
    "country":{"country","interestedcountry","destination","destinationcountry"},
    "visa_type":{"visa","visatype","product","category"},
    "source":{"source","campaign","leadsource","area","state"}
}

def map_headers(headers):
    result={}
    for i,h in enumerate(headers):
        n=normalize_header(h)
        for field,aliases in FIELD_ALIASES.items():
            if n in aliases and field not in result:
                result[field]=i
    return result

def read_excel(path):
    ext=os.path.splitext(path)[1].lower()
    rows=[]
    if ext==".xlsx":
        wb=load_workbook(path,read_only=True,data_only=True)
        ws=wb.active
        rows=list(ws.iter_rows(values_only=True))
    elif ext==".xls":
        book=xlrd.open_workbook(path)
        sh=book.sheet_by_index(0)
        rows=[[sh.cell_value(r,c) for c in range(sh.ncols)] for r in range(sh.nrows)]
    else:
        raise ValueError("Only .xls and .xlsx are supported")
    return rows

@app.route("/leads", methods=["GET","POST"])
@require_roles("MD","GM")
def leads():
    u=current_user()
    con=db()
    if request.method=="POST":
        f=request.files.get("excel_file")
        if not f or not f.filename:
            flash("Select an Excel file","error")
            return redirect(url_for("leads"))
        company=request.form.get("company_code") if u["role"]=="MD" else u["company_code"]
        fn=f"{uuid.uuid4().hex}_{f.filename}"
        path=os.path.join(UPLOAD_DIR,fn); f.save(path)
        try:
            rows=read_excel(path)
            if not rows: raise ValueError("Excel is empty")
            header_idx=0
            best={}
            for idx in range(min(10,len(rows))):
                m=map_headers(rows[idx])
                if len(m)>len(best):
                    best=m; header_idx=idx
            mapping=best
            if "mobile" not in mapping:
                raise ValueError("Mobile/Contact column could not be identified")
            total=new=same=cross=invalid=0
            for raw in rows[header_idx+1:]:
                if not any(str(x).strip() for x in raw if x is not None): continue
                total+=1
                def val(field):
                    i=mapping.get(field)
                    if i is None or i>=len(raw): return ""
                    x=raw[i]
                    return str(x).strip() if x is not None else ""
                mobile=re.sub(r'\D','',val("mobile"))
                dtype=""
                if len(mobile)<8:
                    dtype="invalid"; invalid+=1
                else:
                    same_hit=con.execute("SELECT id FROM leads WHERE company_code=? AND REPLACE(REPLACE(mobile,' ',''),'-','')=? LIMIT 1",(company,mobile)).fetchone()
                    cross_hit=con.execute("SELECT id FROM leads WHERE company_code<>? AND REPLACE(REPLACE(mobile,' ',''),'-','')=? LIMIT 1",(company,mobile)).fetchone()
                    if cross_hit: dtype="cross"; cross+=1
                    elif same_hit: dtype="same"; same+=1
                    else: new+=1
                lead_id=f"{company}-LEAD-{datetime.datetime.now().strftime('%y%m%d')}-{uuid.uuid4().hex[:6].upper()}"
                con.execute("""INSERT INTO leads(lead_id,company_code,client_name,mobile,email,city,country,visa_type,source,duplicate_type,imported_by,imported_at)
                               VALUES(?,?,?,?,?,?,?,?,?,?,?,?)""",
                            (lead_id,company,val("client_name"),mobile,val("email"),val("city"),val("country"),val("visa_type"),
                             val("source"),dtype,u["login_id"],datetime.datetime.now().isoformat(timespec="seconds")))
            con.execute("""INSERT INTO imports(company_code,filename,total_rows,new_rows,same_company_duplicates,cross_company_duplicates,invalid_mobiles,imported_by,imported_at)
                           VALUES(?,?,?,?,?,?,?,?,?)""",(company,f.filename,total,new,same,cross,invalid,u["login_id"],datetime.datetime.now().isoformat(timespec="seconds")))
            con.commit()
            flash(f"Imported {total} rows | New {new} | Same duplicate {same} | Cross duplicate {cross} | Invalid {invalid}","success")
        except Exception as e:
            flash(str(e),"error")
        finally:
            try: os.remove(path)
            except: pass
    company_filter=request.args.get("company","")
    q="SELECT l.*, u.full_name am_name FROM leads l LEFT JOIN users u ON u.id=l.assigned_am"
    params=[]
    clauses=[]
    if u["role"]=="GM":
        clauses.append("l.company_code=?"); params.append(u["company_code"])
    elif company_filter:
        clauses.append("l.company_code=?"); params.append(company_filter)
    if clauses: q+=" WHERE "+" AND ".join(clauses)
    q+=" ORDER BY l.id DESC LIMIT 300"
    rows=con.execute(q,params).fetchall()
    imports=con.execute("SELECT * FROM imports ORDER BY id DESC LIMIT 20").fetchall()
    if u["role"]=="MD":
        am_rows=con.execute("SELECT * FROM users WHERE role='AM' AND active=1 ORDER BY company_code,full_name").fetchall()
    else:
        am_rows=con.execute("SELECT * FROM users WHERE role='AM' AND active=1 AND company_code=? ORDER BY full_name",(u["company_code"],)).fetchall()
    con.close()
    return render_template("leads.html",u=u,rows=rows,imports=imports,am_rows=am_rows)

@app.route("/ams", methods=["GET","POST"])
@require_roles("MD","GM")
def ams():
    u=current_user(); con=db()
    if request.method=="POST":
        company=request.form.get("company_code") if u["role"]=="MD" else u["company_code"]
        name=request.form.get("full_name","").strip()
        pw=request.form.get("password","").strip()
        if not name or len(pw)<6:
            flash("Name and minimum 6-character password required","error")
        else:
            seq=con.execute("SELECT COUNT(*) c FROM users WHERE role='AM' AND company_code=?",(company,)).fetchone()["c"]+1
            lid=f"{company.lower()}-am-{seq:03d}"
            while con.execute("SELECT id FROM users WHERE login_id=?",(lid,)).fetchone():
                seq+=1; lid=f"{company.lower()}-am-{seq:03d}"
            con.execute("INSERT INTO users(login_id,password_hash,full_name,role,company_code) VALUES(?,?,?,?,?)",
                        (lid,generate_password_hash(pw),name,"AM",company))
            con.commit(); flash(f"AM created. Login ID: {lid}","success")
    if u["role"]=="MD":
        rows=con.execute("SELECT * FROM users WHERE role='AM' ORDER BY id DESC").fetchall()
    else:
        rows=con.execute("SELECT * FROM users WHERE role='AM' AND company_code=? ORDER BY id DESC",(u["company_code"],)).fetchall()
    con.close()
    return render_template("ams.html",u=u,rows=rows)



@app.route("/employee-file/<int:document_id>")
@require_roles("MD","GM")
def employee_file(document_id):
    u=current_user(); con=db()
    row=con.execute("""SELECT d.*,e.company_code FROM employee_documents d
                       JOIN employee_master e ON e.id=d.employee_id WHERE d.id=?""",(document_id,)).fetchone()
    con.close()
    if not row:
        flash("Document not found","error"); return redirect(url_for("employees"))
    if u["role"]=="GM" and row["company_code"]!=u["company_code"]:
        flash("Access denied","error"); return redirect(url_for("employees"))
    if row.get("file_data") if isinstance(row,dict) else row["file_data"]:
        data=row["file_data"]
        if isinstance(data,memoryview): data=data.tobytes()
        return send_file(io.BytesIO(data),mimetype=row["mime_type"] or "application/octet-stream",
                         download_name=row["file_name"],as_attachment=False)
    path=os.path.join(APP_DIR,row["stored_path"])
    if not os.path.exists(path):
        flash("File missing from storage","error"); return redirect(url_for("employees"))
    return send_file(path,as_attachment=False)

@app.route("/employee-photo/<int:employee_id>")
@require_roles("MD","GM")
def employee_photo(employee_id):
    u=current_user(); con=db()
    row=con.execute("SELECT company_code,photo_path,photo_data,photo_mime FROM employee_master WHERE id=?",(employee_id,)).fetchone()
    con.close()
    if not row: return ("",404)
    if u["role"]=="GM" and row["company_code"]!=u["company_code"]: return ("",403)
    pdata=row["photo_data"]
    if pdata:
        if isinstance(pdata,memoryview): pdata=pdata.tobytes()
        return send_file(io.BytesIO(pdata),mimetype=row["photo_mime"] or "image/jpeg")
    if not row["photo_path"]: return ("",404)
    path=os.path.join(APP_DIR,row["photo_path"])
    if not os.path.exists(path): return ("",404)
    return send_file(path)


@app.route("/employees", methods=["GET","POST"])
@require_roles("MD","GM")
def employees():
    u=current_user()
    con=db()

    if request.method=="POST":
        company=request.form.get("company_code") if u["role"]=="MD" else u["company_code"]
        full_name=request.form.get("full_name","").strip()
        if not full_name:
            flash("Employee name is required","error")
        else:
            seq=con.execute("SELECT COUNT(*) c FROM employee_master WHERE company_code=?",(company,)).fetchone()["c"]+1
            employee_code=f"{company}-EMP-{seq:04d}"
            while con.execute("SELECT id FROM employee_master WHERE employee_code=?",(employee_code,)).fetchone():
                seq += 1
                employee_code=f"{company}-EMP-{seq:04d}"
            now=datetime.datetime.now().isoformat(timespec="seconds")
            cur=con.execute("""INSERT INTO employee_master(
                employee_code,company_code,full_name,father_spouse_name,designation,department,
                mobile,alternate_mobile,email,date_of_birth,joining_date,address,city,state,pin_code,
                emergency_contact_name,emergency_contact_number,aadhaar_last4,pan_number,bank_name,
                account_last4,ifsc_code,salary_reference,employment_status,remarks,photo_path,created_by,created_at,updated_at
            ) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",(
                employee_code,company,full_name,request.form.get("father_spouse_name",""),
                request.form.get("designation",""),request.form.get("department",""),
                request.form.get("mobile",""),request.form.get("alternate_mobile",""),
                request.form.get("email",""),request.form.get("date_of_birth",""),
                request.form.get("joining_date",""),request.form.get("address",""),
                request.form.get("city",""),request.form.get("state",""),request.form.get("pin_code",""),
                request.form.get("emergency_contact_name",""),request.form.get("emergency_contact_number",""),
                request.form.get("aadhaar_last4",""),request.form.get("pan_number",""),
                request.form.get("bank_name",""),request.form.get("account_last4",""),
                request.form.get("ifsc_code",""),request.form.get("salary_reference",""),
                request.form.get("employment_status","Active"),request.form.get("remarks",""),
                "",u["login_id"],now,now
            ))
            employee_id=cur.lastrowid

            photo=request.files.get("photograph")
            if photo and photo.filename and allowed_file(photo.filename,ALLOWED_PHOTO_EXTENSIONS):
                mime=photo.mimetype or mimetypes.guess_type(photo.filename)[0] or "image/jpeg"
                if IS_POSTGRES or os.environ.get("STORE_FILES_IN_DB","1")=="1":
                    con.execute("UPDATE employee_master SET photo_path='DB',photo_data=?,photo_mime=? WHERE id=?",
                                (photo.read(),mime,employee_id))
                else:
                    employee_dir=os.path.join(EMPLOYEE_UPLOAD_DIR,employee_code); os.makedirs(employee_dir,exist_ok=True)
                    ext=photo.filename.rsplit(".",1)[1].lower(); full_path=os.path.join(employee_dir,f"photograph.{ext}")
                    photo.save(full_path); rel_path=os.path.relpath(full_path,APP_DIR).replace("\\","/")
                    con.execute("UPDATE employee_master SET photo_path=? WHERE id=?",(rel_path,employee_id))

            con.commit()
            flash(f"Employee saved successfully. Employee Code: {employee_code}","success")

    if u["role"]=="MD":
        rows=con.execute("SELECT * FROM employee_master ORDER BY id DESC").fetchall()
    else:
        rows=con.execute("SELECT * FROM employee_master WHERE company_code=? ORDER BY id DESC",(u["company_code"],)).fetchall()
    doc_counts={r["employee_id"]:r["c"] for r in con.execute("SELECT employee_id,COUNT(*) c FROM employee_documents GROUP BY employee_id").fetchall()}
    con.close()
    return render_template("employees.html",u=u,rows=rows,doc_counts=doc_counts)



def performance_range(period):
    today=datetime.date.today()
    if period=="week":
        start=today-datetime.timedelta(days=today.weekday())
        end=start+datetime.timedelta(days=6)
        label=f"Weekly Report ({start.strftime('%d-%b-%Y')} to {end.strftime('%d-%b-%Y')})"
    elif period=="year":
        start=datetime.date(today.year,1,1)
        end=datetime.date(today.year,12,31)
        label=f"Yearly Report {today.year}"
    else:
        start=datetime.date(today.year,today.month,1)
        if today.month==12:
            next_month=datetime.date(today.year+1,1,1)
        else:
            next_month=datetime.date(today.year,today.month+1,1)
        end=next_month-datetime.timedelta(days=1)
        label=today.strftime("Monthly Report - %B %Y")
        period="month"
    return period,start.isoformat(),end.isoformat(),label


@app.route("/performance")
@require_roles("MD","GM")
def performance_directory():
    u=current_user()
    con=db()
    if u["role"]=="MD":
        employees=con.execute("SELECT * FROM employee_master WHERE employment_status='Active' ORDER BY company_code,full_name").fetchall()
    else:
        employees=con.execute("SELECT * FROM employee_master WHERE company_code=? AND employment_status='Active' ORDER BY full_name",(u["company_code"],)).fetchall()

    period,start_date,end_date,label=performance_range(request.args.get("period","month"))
    raw=[]
    for e in employees:
        s=con.execute("""SELECT
              COALESCE(SUM(leads_assigned),0) leads_assigned,
              COALESCE(SUM(calls_done),0) calls_done,
              COALESCE(SUM(followups_done),0) followups_done,
              COALESCE(SUM(office_visits),0) office_visits,
              COALESCE(SUM(interested_clients),0) interested_clients,
              COALESCE(SUM(enrollments),0) enrollments,
              COALESCE(SUM(collection_amount),0) collection_amount,
              COALESCE(SUM(target_amount),0) target_amount,
              COALESCE(AVG(NULLIF(manager_rating,0)),0) avg_rating
            FROM employee_performance
            WHERE employee_id=? AND performance_date BETWEEN ? AND ?""",
            (e["id"],start_date,end_date)).fetchone()
        d=dict(s)
        # Enrollment and revenue ranking are automatically taken from Client Cases.
        # Calls/follow-ups/rating continue to come from daily performance entries.
        case_stats=con.execute("""SELECT COUNT(*) enrollments, COALESCE(SUM(total_received),0) revenue
                                  FROM client_cases
                                  WHERE assigned_employee_id=? AND enrollment_date BETWEEN ? AND ?""",
                               (e["id"],start_date,end_date)).fetchone()
        d["enrollments"]=case_stats["enrollments"]
        d["collection_amount"]=case_stats["revenue"]
        d["employee"]=e
        raw.append(d)

    max_enrollments=max([r["enrollments"] or 0 for r in raw], default=0)
    max_revenue=max([r["collection_amount"] or 0 for r in raw], default=0)
    total_enrollments=sum(r["enrollments"] or 0 for r in raw)
    total_revenue=sum(r["collection_amount"] or 0 for r in raw)
    summaries={}
    ranked=[]

    for r in raw:
        e=r["employee"]
        enrollments=r["enrollments"] or 0
        revenue=r["collection_amount"] or 0
        target=r["target_amount"] or 0
        rating=r["avg_rating"] or 0
        leads=r["leads_assigned"] or 0

        enrollment_index=(enrollments/max_enrollments*100) if max_enrollments else 0
        revenue_index=(revenue/max_revenue*100) if max_revenue else 0
        target_achievement=(revenue/target*100) if target else 0
        conversion=(enrollments/leads*100) if leads else 0

        score=(min(enrollment_index,100)*0.40 +
               min(revenue_index,100)*0.40 +
               min(target_achievement,100)*0.10 +
               min(rating*20,100)*0.10)

        r.update({
            "enrollment_index":enrollment_index,
            "revenue_index":revenue_index,
            "target_achievement":target_achievement,
            "conversion":conversion,
            "score":score
        })
        summaries[e["id"]]=r
        ranked.append({
            "id":e["id"],
            "employee_code":e["employee_code"],
            "full_name":e["full_name"],
            "company_code":e["company_code"],
            "designation":e["designation"],
            "photo_path":e["photo_path"],
            "score":score,
            "enrollments":enrollments,
            "revenue":revenue,
            "target_achievement":target_achievement,
            "conversion":conversion,
            "rating":rating
        })

    ranked.sort(key=lambda x:(x["score"],x["revenue"],x["enrollments"]), reverse=True)
    for idx,r in enumerate(ranked,1):
        r["rank"]=idx
    top3=ranked[:3]

    active_employees=len(employees)
    avg_enrollments=(total_enrollments/active_employees) if active_employees else 0
    avg_revenue=(total_revenue/active_employees) if active_employees else 0
    top_performer=top3[0] if top3 else None

    # Build month-by-month trend for selected calendar year using the same ranking formula.
    today=datetime.date.today()
    year=today.year
    trend=[]
    # Take up to 3 current top performers and show their monthly performance scores.
    trend_employees=top3
    for month in range(1,13):
        month_start=datetime.date(year,month,1)
        month_end=(datetime.date(year+1,1,1)-datetime.timedelta(days=1)) if month==12 else (datetime.date(year,month+1,1)-datetime.timedelta(days=1))
        month_raw=[]
        for emp in trend_employees:
            s=con.execute("""SELECT
                  COALESCE(SUM(enrollments),0) enrollments,
                  COALESCE(SUM(collection_amount),0) collection_amount,
                  COALESCE(SUM(target_amount),0) target_amount,
                  COALESCE(AVG(NULLIF(manager_rating,0)),0) avg_rating
                FROM employee_performance
                WHERE employee_id=? AND performance_date BETWEEN ? AND ?""",
                (emp["id"],month_start.isoformat(),month_end.isoformat())).fetchone()
            md=dict(s)
            case_stats=con.execute("""SELECT COUNT(*) enrollments, COALESCE(SUM(total_received),0) revenue
                                      FROM client_cases WHERE assigned_employee_id=? AND enrollment_date BETWEEN ? AND ?""",
                                   (emp["id"],month_start.isoformat(),month_end.isoformat())).fetchone()
            md["enrollments"]=case_stats["enrollments"]
            md["collection_amount"]=case_stats["revenue"]
            month_raw.append({"id":emp["id"],**md})
        month_max_enroll=max([m["enrollments"] or 0 for m in month_raw],default=0)
        month_max_revenue=max([m["collection_amount"] or 0 for m in month_raw],default=0)
        item={"month":month_start.strftime("%b")}
        for idx,m in enumerate(month_raw,1):
            enrollment_index=(m["enrollments"]/month_max_enroll*100) if month_max_enroll else 0
            revenue_index=(m["collection_amount"]/month_max_revenue*100) if month_max_revenue else 0
            target_achievement=(m["collection_amount"]/m["target_amount"]*100) if m["target_amount"] else 0
            rating=m["avg_rating"] or 0
            mscore=(min(enrollment_index,100)*0.40 +
                    min(revenue_index,100)*0.40 +
                    min(target_achievement,100)*0.10 +
                    min(rating*20,100)*0.10)
            item[f"score{idx}"]=round(mscore,1)
        trend.append(item)

    con.close()
    return render_template("performance_directory.html",u=u,employees=employees,summaries=summaries,
                           ranked=ranked,top3=top3,period=period,label=label,start_date=start_date,end_date=end_date,
                           total_enrollments=total_enrollments,total_revenue=total_revenue,
                           avg_enrollments=avg_enrollments,avg_revenue=avg_revenue,
                           active_employees=active_employees,top_performer=top_performer,trend=trend)


@app.route("/employees/<int:employee_id>/performance", methods=["GET","POST"])
@require_roles("MD","GM")
def employee_performance(employee_id):
    u=current_user()
    con=db()
    employee=con.execute("SELECT * FROM employee_master WHERE id=?",(employee_id,)).fetchone()
    if not employee:
        con.close()
        flash("Employee record not found","error")
        return redirect(url_for("employees"))
    if u["role"]=="GM" and employee["company_code"]!=u["company_code"]:
        con.close()
        flash("Access denied","error")
        return redirect(url_for("employees"))

    if request.method=="POST":
        pdate=request.form.get("performance_date") or datetime.date.today().isoformat()
        try:
            rating=float(request.form.get("manager_rating","0") or 0)
        except:
            rating=0
        rating=max(0,min(5,rating))
        con.execute("""INSERT INTO employee_performance(
            employee_id,performance_date,leads_assigned,calls_done,followups_done,office_visits,
            interested_clients,enrollments,collection_amount,target_amount,attendance_status,
            manager_rating,remarks,created_by,created_at
        ) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",(
            employee_id,pdate,
            int(request.form.get("leads_assigned","0") or 0),
            int(request.form.get("calls_done","0") or 0),
            int(request.form.get("followups_done","0") or 0),
            int(request.form.get("office_visits","0") or 0),
            int(request.form.get("interested_clients","0") or 0),
            int(request.form.get("enrollments","0") or 0),
            float(request.form.get("collection_amount","0") or 0),
            float(request.form.get("target_amount","0") or 0),
            request.form.get("attendance_status","Present"),
            rating,request.form.get("remarks",""),u["login_id"],
            datetime.datetime.now().isoformat(timespec="seconds")
        ))
        con.commit()
        flash("Performance entry saved","success")
        return redirect(url_for("employee_performance",employee_id=employee_id,period=request.args.get("period","month")))

    period,start_date,end_date,label=performance_range(request.args.get("period","month"))
    rows=con.execute("""SELECT * FROM employee_performance
                        WHERE employee_id=? AND performance_date BETWEEN ? AND ?
                        ORDER BY performance_date DESC,id DESC""",
                     (employee_id,start_date,end_date)).fetchall()

    totals=dict(con.execute("""SELECT
          COALESCE(SUM(leads_assigned),0) leads_assigned,
          COALESCE(SUM(calls_done),0) calls_done,
          COALESCE(SUM(followups_done),0) followups_done,
          COALESCE(SUM(office_visits),0) office_visits,
          COALESCE(SUM(interested_clients),0) interested_clients,
          COALESCE(SUM(enrollments),0) enrollments,
          COALESCE(SUM(collection_amount),0) collection_amount,
          COALESCE(SUM(target_amount),0) target_amount,
          COALESCE(AVG(NULLIF(manager_rating,0)),0) avg_rating,
          COUNT(*) entries
        FROM employee_performance
        WHERE employee_id=? AND performance_date BETWEEN ? AND ?""",
        (employee_id,start_date,end_date)).fetchone())
    case_stats=con.execute("""SELECT COUNT(*) enrollments,COALESCE(SUM(total_received),0) revenue
                              FROM client_cases WHERE assigned_employee_id=? AND enrollment_date BETWEEN ? AND ?""",
                           (employee_id,start_date,end_date)).fetchone()
    totals["enrollments"]=case_stats["enrollments"]
    totals["collection_amount"]=case_stats["revenue"]
    con.close()

    conversion=(totals["enrollments"]/totals["leads_assigned"]*100) if totals["leads_assigned"] else 0
    achievement=(totals["collection_amount"]/totals["target_amount"]*100) if totals["target_amount"] else 0

    share_text=(
        f"SCIC | WWIC Employee Performance\\n"
        f"Employee: {employee['full_name']} ({employee['employee_code']})\\n"
        f"Company: {employee['company_code']}\\n"
        f"{label}\\n\\n"
        f"Leads Assigned: {totals['leads_assigned']}\\n"
        f"Calls Done: {totals['calls_done']}\\n"
        f"Follow-ups: {totals['followups_done']}\\n"
        f"Office Visits: {totals['office_visits']}\\n"
        f"Interested Clients: {totals['interested_clients']}\\n"
        f"Enrollments: {totals['enrollments']}\\n"
        f"Conversion: {conversion:.1f}%\\n"
        f"Collection: ₹{totals['collection_amount']:,.0f}\\n"
        f"Target: ₹{totals['target_amount']:,.0f}\\n"
        f"Target Achievement: {achievement:.1f}%\\n"
        f"Manager Rating: {totals['avg_rating']:.1f}/5"
    )
    whatsapp_url="https://wa.me/?text="+quote(share_text)

    return render_template("employee_performance.html",u=u,employee=employee,rows=rows,
                           totals=totals,period=period,label=label,start_date=start_date,end_date=end_date,
                           conversion=conversion,achievement=achievement,share_text=share_text,
                           whatsapp_url=whatsapp_url)


@app.route("/employees/<int:employee_id>/documents", methods=["GET","POST"])
@require_roles("MD","GM")
def employee_documents(employee_id):
    u=current_user()
    con=db()
    employee=con.execute("SELECT * FROM employee_master WHERE id=?",(employee_id,)).fetchone()
    if not employee:
        con.close()
        flash("Employee record not found","error")
        return redirect(url_for("employees"))
    if u["role"]=="GM" and employee["company_code"]!=u["company_code"]:
        con.close()
        flash("Access denied","error")
        return redirect(url_for("employees"))

    if request.method=="POST":
        f=request.files.get("document_file")
        doc_type=request.form.get("document_type","Other").strip()
        doc_name=request.form.get("document_name","").strip()
        if not f or not f.filename:
            flash("Select a document file","error")
        elif not allowed_file(f.filename,ALLOWED_DOC_EXTENSIONS):
            flash("Allowed formats: PDF, JPG, JPEG, PNG, WEBP, DOC, DOCX","error")
        else:
            safe=secure_filename(f.filename)
            mime=f.mimetype or mimetypes.guess_type(safe)[0] or "application/octet-stream"
            now=datetime.datetime.now().isoformat(timespec="seconds")
            if IS_POSTGRES or os.environ.get("STORE_FILES_IN_DB","1")=="1":
                con.execute("""INSERT INTO employee_documents(employee_id,document_type,document_name,file_name,stored_path,file_data,mime_type,uploaded_by,uploaded_at)
                               VALUES(?,?,?,?,?,?,?,?,?)""",
                            (employee_id,doc_type,doc_name,safe,"DB",f.read(),mime,u["login_id"],now))
            else:
                emp_dir=os.path.join(EMPLOYEE_UPLOAD_DIR,employee["employee_code"],"documents"); os.makedirs(emp_dir,exist_ok=True)
                stamp=datetime.datetime.now().strftime("%Y%m%d_%H%M%S_%f"); full_path=os.path.join(emp_dir,f"{stamp}_{safe}")
                f.save(full_path); rel_path=os.path.relpath(full_path,APP_DIR).replace("\\","/")
                con.execute("""INSERT INTO employee_documents(employee_id,document_type,document_name,file_name,stored_path,uploaded_by,uploaded_at)
                               VALUES(?,?,?,?,?,?,?)""",(employee_id,doc_type,doc_name,safe,rel_path,u["login_id"],now))
            con.commit()
            flash("Employee document uploaded","success")

    docs=con.execute("SELECT * FROM employee_documents WHERE employee_id=? ORDER BY id DESC",(employee_id,)).fetchall()
    con.close()
    return render_template("employee_documents.html",u=u,employee=employee,docs=docs)


@app.route("/employees/<int:employee_id>/edit", methods=["GET","POST"])
@require_roles("MD","GM")
def edit_employee(employee_id):
    u=current_user()
    con=db()
    row=con.execute("SELECT * FROM employee_master WHERE id=?",(employee_id,)).fetchone()
    if not row:
        con.close()
        flash("Employee record not found","error")
        return redirect(url_for("employees"))
    if u["role"]=="GM" and row["company_code"]!=u["company_code"]:
        con.close()
        flash("Access denied","error")
        return redirect(url_for("employees"))

    if request.method=="POST":
        company=row["company_code"] if u["role"]=="GM" else request.form.get("company_code",row["company_code"])
        now=datetime.datetime.now().isoformat(timespec="seconds")
        con.execute("""UPDATE employee_master SET
            company_code=?,full_name=?,father_spouse_name=?,designation=?,department=?,mobile=?,
            alternate_mobile=?,email=?,date_of_birth=?,joining_date=?,address=?,city=?,state=?,pin_code=?,
            emergency_contact_name=?,emergency_contact_number=?,aadhaar_last4=?,pan_number=?,bank_name=?,
            account_last4=?,ifsc_code=?,salary_reference=?,employment_status=?,remarks=?,updated_at=?
            WHERE id=?""",(
            company,request.form.get("full_name","").strip(),request.form.get("father_spouse_name",""),
            request.form.get("designation",""),request.form.get("department",""),
            request.form.get("mobile",""),request.form.get("alternate_mobile",""),
            request.form.get("email",""),request.form.get("date_of_birth",""),
            request.form.get("joining_date",""),request.form.get("address",""),
            request.form.get("city",""),request.form.get("state",""),request.form.get("pin_code",""),
            request.form.get("emergency_contact_name",""),request.form.get("emergency_contact_number",""),
            request.form.get("aadhaar_last4",""),request.form.get("pan_number",""),
            request.form.get("bank_name",""),request.form.get("account_last4",""),
            request.form.get("ifsc_code",""),request.form.get("salary_reference",""),
            request.form.get("employment_status","Active"),request.form.get("remarks",""),
            now,employee_id
        ))
        photo=request.files.get("photograph")
        if photo and photo.filename:
            if allowed_file(photo.filename,ALLOWED_PHOTO_EXTENSIONS):
                mime=photo.mimetype or mimetypes.guess_type(photo.filename)[0] or "image/jpeg"
                if IS_POSTGRES or os.environ.get("STORE_FILES_IN_DB","1")=="1":
                    con.execute("UPDATE employee_master SET photo_path='DB',photo_data=?,photo_mime=?,updated_at=? WHERE id=?",
                                (photo.read(),mime,now,employee_id))
                else:
                    employee_dir=os.path.join(EMPLOYEE_UPLOAD_DIR,row["employee_code"]); os.makedirs(employee_dir,exist_ok=True)
                    ext=photo.filename.rsplit(".",1)[1].lower(); full_path=os.path.join(employee_dir,f"photograph.{ext}")
                    photo.save(full_path); rel_path=os.path.relpath(full_path,APP_DIR).replace("\\","/")
                    con.execute("UPDATE employee_master SET photo_path=?,updated_at=? WHERE id=?",(rel_path,now,employee_id))
            else:
                flash("Photograph must be JPG, JPEG, PNG or WEBP","error")
        con.commit()
        con.close()
        flash("Employee information updated","success")
        return redirect(url_for("employees"))

    con.close()
    return render_template("employee_edit.html",u=u,row=row)



@app.route("/ams/<int:user_id>/toggle", methods=["POST"])
@require_roles("MD","GM")
def toggle_am(user_id):
    u=current_user(); con=db()
    am=con.execute("SELECT * FROM users WHERE id=? AND role='AM'",(user_id,)).fetchone()
    if not am or (u["role"]=="GM" and am["company_code"]!=u["company_code"]):
        con.close(); flash("AM not found or access denied","error"); return redirect(url_for("ams"))
    new_active=0 if am["active"] else 1
    con.execute("UPDATE users SET active=? WHERE id=?",(new_active,user_id)); con.commit(); con.close()
    flash("AM portal activated" if new_active else "AM portal deactivated. Data has been preserved.","success")
    return redirect(url_for("ams"))

@app.route("/ams/handover", methods=["POST"])
@require_roles("MD","GM")
def handover_am():
    u=current_user(); con=db()
    old_id=int(request.form.get("old_am_id","0") or 0); new_id=int(request.form.get("new_am_id","0") or 0)
    old=con.execute("SELECT * FROM users WHERE id=? AND role='AM'",(old_id,)).fetchone()
    new=con.execute("SELECT * FROM users WHERE id=? AND role='AM' AND active=1",(new_id,)).fetchone()
    if not old or not new or old_id==new_id or old["company_code"]!=new["company_code"] or (u["role"]=="GM" and old["company_code"]!=u["company_code"]):
        con.close(); flash("Select two valid AMs from the same company","error"); return redirect(url_for("ams"))
    count=con.execute("SELECT COUNT(*) c FROM leads WHERE assigned_am=?",(old_id,)).fetchone()["c"]
    now=datetime.datetime.now().isoformat(timespec="seconds")
    con.execute("UPDATE leads SET assigned_am=?,assigned_at=?,assigned_by=? WHERE assigned_am=?",(new_id,now,u["login_id"],old_id))
    con.execute("UPDATE users SET active=0 WHERE id=?",(old_id,))
    con.execute("INSERT INTO allocation_history(company_code,am_user_id,quantity,allocated_by,allocated_at) VALUES(?,?,?,?,?)",(new["company_code"],new_id,count,u["login_id"]+" HANDOVER",now))
    con.commit(); con.close(); flash(f"Handover complete: {count} leads transferred. Old AM portal deactivated.","success")
    return redirect(url_for("ams"))

@app.route("/cases", methods=["GET","POST"])
@require_roles("MD","GM")
def cases():
    u=current_user(); con=db()
    if request.method=="POST":
        company=request.form.get("company_code") if u["role"]=="MD" else u["company_code"]
        cname=request.form.get("client_name","").strip()
        if not cname:
            flash("Client name is required","error")
        else:
            seq=con.execute("SELECT COUNT(*) c FROM client_cases WHERE company_code=?",(company,)).fetchone()["c"]+1
            case_id=f"{company}-CASE-{datetime.date.today().strftime('%y%m%d')}-{seq:04d}"
            while con.execute("SELECT id FROM client_cases WHERE case_id=?",(case_id,)).fetchone():
                seq+=1; case_id=f"{company}-CASE-{datetime.date.today().strftime('%y%m%d')}-{seq:04d}"
            b=float(request.form.get("booking_amount","0") or 0); s2=float(request.form.get("second_payment","0") or 0)
            total=b+s2; now=datetime.datetime.now().isoformat(timespec="seconds")
            con.execute("""INSERT INTO client_cases(case_id,company_code,client_name,mobile,country,visa_type,enrollment_date,booking_amount,second_payment,total_received,payment_status,filing_status,assigned_employee_id,remarks,created_by,created_at,updated_at) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                        (case_id,company,cname,request.form.get("mobile",""),request.form.get("country",""),request.form.get("visa_type",""),request.form.get("enrollment_date",datetime.date.today().isoformat()),b,s2,total,request.form.get("payment_status","Pending"),request.form.get("filing_status","Documents Pending"),request.form.get("assigned_employee_id") or None,request.form.get("remarks",""),u["login_id"],now,now))
            con.commit(); flash(f"Client case created: {case_id}","success")
    if u["role"]=="MD":
        rows=con.execute("SELECT c.*,e.full_name employee_name FROM client_cases c LEFT JOIN employee_master e ON e.id=c.assigned_employee_id ORDER BY c.id DESC").fetchall()
        staff=con.execute("SELECT id,full_name,company_code FROM employee_master WHERE employment_status='Active' ORDER BY company_code,full_name").fetchall()
    else:
        rows=con.execute("SELECT c.*,e.full_name employee_name FROM client_cases c LEFT JOIN employee_master e ON e.id=c.assigned_employee_id WHERE c.company_code=? ORDER BY c.id DESC",(u["company_code"],)).fetchall()
        staff=con.execute("SELECT id,full_name,company_code FROM employee_master WHERE employment_status='Active' AND company_code=? ORDER BY full_name",(u["company_code"],)).fetchall()
    con.close(); return render_template("cases.html",u=u,rows=rows,staff=staff)

@app.route("/cases/<int:case_id>/update", methods=["POST"])
@require_roles("MD","GM")
def update_case(case_id):
    u=current_user(); con=db(); row=con.execute("SELECT * FROM client_cases WHERE id=?",(case_id,)).fetchone()
    if not row or (u["role"]=="GM" and row["company_code"]!=u["company_code"]):
        con.close(); flash("Case not found or access denied","error"); return redirect(url_for("cases"))
    b=float(request.form.get("booking_amount",row["booking_amount"]) or 0); s2=float(request.form.get("second_payment",row["second_payment"]) or 0)
    con.execute("""UPDATE client_cases SET booking_amount=?,second_payment=?,total_received=?,payment_status=?,filing_status=?,remarks=?,updated_at=? WHERE id=?""",
                (b,s2,b+s2,request.form.get("payment_status",row["payment_status"]),request.form.get("filing_status",row["filing_status"]),request.form.get("remarks",row["remarks"]),datetime.datetime.now().isoformat(timespec="seconds"),case_id))
    con.commit(); con.close(); flash("Case updated","success"); return redirect(url_for("cases"))

@app.route("/allocate", methods=["POST"])
@require_roles("MD","GM")
def allocate():
    u=current_user(); con=db()
    am_id=request.form.get("am_id")
    qty=int(request.form.get("qty","0") or 0)
    am=con.execute("SELECT * FROM users WHERE id=? AND role='AM' AND active=1",(am_id,)).fetchone()
    if not am:
        flash("Select a valid AM","error")
    else:
        if u["role"]=="GM" and am["company_code"]!=u["company_code"]:
            flash("Cannot allocate outside your company","error")
        else:
            ids=[r["id"] for r in con.execute("SELECT id FROM leads WHERE company_code=? AND assigned_am IS NULL ORDER BY id LIMIT ?",(am["company_code"],qty)).fetchall()]
            now=datetime.datetime.now().isoformat(timespec="seconds")
            for lid in ids:
                con.execute("UPDATE leads SET assigned_am=?,assigned_at=?,assigned_by=? WHERE id=?",(am["id"],now,u["login_id"],lid))
            con.execute("INSERT INTO allocation_history(company_code,am_user_id,quantity,allocated_by,allocated_at) VALUES(?,?,?,?,?)",(am["company_code"],am["id"],len(ids),u["login_id"],now))
            con.commit()
            flash(f"{len(ids)} leads allocated to {am['full_name']}","success")
    con.close()
    return redirect(url_for("leads"))

@app.route("/my-leads", methods=["GET","POST"])
@require_roles("AM")
def my_leads():
    u=current_user(); con=db()
    if request.method=="POST":
        lid=request.form.get("lead_id")
        con.execute("UPDATE leads SET status=?,followup_date=?,remarks=? WHERE id=? AND assigned_am=?",
                    (request.form.get("status"),request.form.get("followup_date"),request.form.get("remarks"),lid,u["id"]))
        con.commit(); flash("Lead updated","success")
    rows=con.execute("SELECT * FROM leads WHERE assigned_am=? ORDER BY id DESC",(u["id"],)).fetchall()
    con.close()
    return render_template("my_leads.html",u=u,rows=rows)

@app.context_processor
def inject_user():
    return dict(current_user=current_user())

if __name__=="__main__":
    import traceback
    try:
        init_db()
        port=int(os.environ.get("PORT","5050"))
        print(f"GAUR CRM ready | DB={'PostgreSQL' if IS_POSTGRES else 'SQLite'} | http://127.0.0.1:{port}",flush=True)
        from waitress import serve
        serve(app,host="0.0.0.0",port=port,threads=8)
    except Exception:
        print("\nGAUR CRM STARTUP ERROR:\n",flush=True); traceback.print_exc(); raise
